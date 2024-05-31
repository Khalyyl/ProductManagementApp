import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl import Workbook

class ProductManagementApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Product Management")
        self.master.geometry("600x400")
        self.master.configure(bg="white")
        self.master.resizable(False, False)

        self.style = ttk.Style()
        self.style.configure("TFrame", background="white")
        self.style.configure("TButton", font=("Arial", 10))

        # Custom styles for colored buttons with yellow background and black text
        self.style.configure("Add.TButton", background="#28a745", foreground="black")
        self.style.configure("Remove.TButton", background="#dc3545", foreground="black")
        self.style.configure("Update.TButton", background="#ffc107", foreground="black")
        self.style.configure("Search.TButton", background="#007BFF", foreground="black")

        self.create_widgets()

        # Load existing data or create a new workbook
        self.load_data()

        # Display all products in the listbox
        self.populate_product_list()

    def create_widgets(self):
        frame = ttk.Frame(self.master, padding="10", style="TFrame")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Labels
        ttk.Label(frame, text="Product Name:", style="TLabel").grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
        ttk.Label(frame, text="Product Price:", style="TLabel").grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
        ttk.Label(frame, text="Quantity Available:", style="TLabel").grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)

        # Entry widgets
        self.entry_name = ttk.Entry(frame, style="TEntry")
        self.entry_price = ttk.Entry(frame, style="TEntry")
        self.entry_quantity = ttk.Entry(frame, style="TEntry")

        # Listbox for displaying products
        self.product_listbox = tk.Listbox(frame, selectmode=tk.SINGLE, height=10, font=("Arial", 10))
        self.product_listbox.bind("<<ListboxSelect>>", self.on_product_select)

        # Search Entry and Button
        self.entry_search = ttk.Entry(frame, style="TEntry")
        self.btn_search = ttk.Button(frame, text="Search", command=self.search_product, style="Search.TButton")

        # Buttons
        self.btn_add = ttk.Button(frame, text="Add Product", command=self.add_product, style="Add.TButton")
        self.btn_remove = ttk.Button(frame, text="Remove Product", command=self.remove_product, style="Remove.TButton")
        self.btn_update = ttk.Button(frame, text="Update Product", command=self.update_product, style="Update.TButton")

        # Placement of widgets
        self.entry_name.grid(row=0, column=1, padx=10, pady=5, sticky=tk.W)
        self.entry_price.grid(row=1, column=1, padx=10, pady=5, sticky=tk.W)
        self.entry_quantity.grid(row=2, column=1, padx=10, pady=5, sticky=tk.W)

        self.product_listbox.grid(row=3, column=0, columnspan=2, padx=10, pady=5, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.entry_search.grid(row=4, column=0, padx=10, pady=5, sticky=tk.W)
        self.btn_search.grid(row=4, column=1, padx=10, pady=5, sticky=tk.W)

        self.btn_add.grid(row=5, column=0, pady=10, sticky=tk.W)
        self.btn_remove.grid(row=5, column=1, pady=10, sticky=tk.W)
        self.btn_update.grid(row=5, column=2, pady=10, sticky=tk.W)

    def load_data(self):
        try:
            self.workbook = openpyxl.load_workbook("products.xlsx")
            self.sheet = self.workbook.active
        except FileNotFoundError:
            # Create a new workbook if the file doesn't exist
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.append(["Product Name", "Price", "Quantity"])

    def populate_product_list(self):
        # Clear existing items in the listbox
        self.product_listbox.delete(0, tk.END)

        # Retrieve product names and add them to the listbox
        product_names = [row[0].value for row in self.sheet.iter_rows(min_row=2, max_col=1)]
        for name in product_names:
            self.product_listbox.insert(tk.END, name)

    def on_product_select(self, event):
        # When a product is selected in the listbox, populate the entry fields with its details
        selected_index = self.product_listbox.curselection()
        if selected_index:
            selected_name = self.product_listbox.get(selected_index)
            selected_product = self.get_product_details(selected_name)
            if selected_product:
                self.entry_name.delete(0, tk.END)
                self.entry_price.delete(0, tk.END)
                self.entry_quantity.delete(0, tk.END)
                self.entry_name.insert(0, selected_product[0].value)
                self.entry_price.insert(0, selected_product[1].value)
                self.entry_quantity.insert(0, selected_product[2].value)

    def get_product_details(self, name):
        # Retrieve details of the selected product
        for row in self.sheet.iter_rows(min_row=2, max_col=3):
            if row[0].value == name:
                return row
        return None

    def add_product(self):
        name = self.entry_name.get()
        price = self.entry_price.get()
        quantity = self.entry_quantity.get()

        if name and price and quantity:
            self.sheet.append([name, price, quantity])
            self.workbook.save("products.xlsx")
            messagebox.showinfo("Success", "Product added successfully!")
            self.clear_entries()
            self.populate_product_list()
        else:
            messagebox.showerror("Error", "Please fill in all fields.")

    def remove_product(self):
        name_to_remove = self.entry_name.get()

        for row in self.sheet.iter_rows(min_row=2, max_col=1):
            if row[0].value == name_to_remove:
                self.sheet.delete_rows(row[0].row)
                self.workbook.save("products.xlsx")
                messagebox.showinfo("Success", "Product removed successfully!")
                self.clear_entries()
                self.populate_product_list()
                return

        messagebox.showerror("Error", "Product not found.")

    def update_product(self):
        name_to_update = self.product_listbox.get(tk.ACTIVE)
        if not name_to_update:
            messagebox.showerror("Error", "Please select a product from the list.")
            return

        for row in self.sheet.iter_rows(min_row=2, max_col=3):
            if row[0].value == name_to_update:
                updated_row = [self.entry_name.get(), self.entry_price.get(), self.entry_quantity.get()]
                for cell, value in zip(row, updated_row):
                    cell.value = value
                self.workbook.save("products.xlsx")
                messagebox.showinfo("Success", "Product updated successfully!")
                self.clear_entries()
                self.populate_product_list()
                return

        messagebox.showerror("Error", "Product not found.")

    def search_product(self):
        search_term = self.entry_search.get().lower()
        if not search_term:
            messagebox.showerror("Error", "Please enter a search term.")
            return

        matching_products = []
        for row in self.sheet.iter_rows(min_row=2, max_col=1):
            if search_term in str(row[0].value).lower():
                matching_products.append(row[0].value)

        if matching_products:
            self.product_listbox.delete(0, tk.END)
            for product in matching_products:
                self.product_listbox.insert(tk.END, product)
        else:
            messagebox.showinfo("No Match", "No products match the search term.")

    def clear_entries(self):
        self.entry_name.delete(0, tk.END)
        self.entry_price.delete(0, tk.END)
        self.entry_quantity.delete(0, tk.END)
        self.entry_search.delete(0, tk.END)

if __name__ == "__main__":
    root = tk.Tk()
    app = ProductManagementApp(root)
    root.mainloop()
